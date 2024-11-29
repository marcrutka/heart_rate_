from confluent_kafka import Consumer
import json
import logging
import pandas as pd
import time
from collections import deque
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
import numpy as np
import requests

SERVER = 'localhost:9092'
KAFKA_TOPIC = 'heartrate'
GROUP_ID = 'heartrate-consumer'
WINDOW_SIZE = 10  # Number of messages to consider for trend analysis


ALERT_THRESHOLDS = {
    'resting': 90,
    'walking': 130,
    'running': 145
}

def create_consumer():
    try:
        consumer = Consumer({
            'bootstrap.servers': SERVER,
            'group.id': GROUP_ID,
            'auto.offset.reset': 'latest'  
        })
        consumer.subscribe([KAFKA_TOPIC])
    except Exception as e:
        logging.exception("Unable to create consumer")
        consumer = None
    return consumer

def process_message(message):
    data = json.loads(message.value().decode('utf-8'))
    time_received = data['time']
    full_date = time_received
    date, time = time_received.split('T')
    hour, minute, second = time.split(':')
    time_received = f"{hour}:{minute}:{second.split('.')[0]}"  
    heart_rate = data['heart_rate']
    activity_type = data.get('activity_type', 'resting') 
    return date, time_received, heart_rate, activity_type, full_date

def add_line_chart_to_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    chart = LineChart()
    chart.title = "Heart Rate Over Time"
    chart.style = 13
    chart.y_axis.title = 'Heart Rate'
    chart.x_axis.title = 'Time'

    data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
    cats = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "K8")  
    wb.save(filename)


consumer = create_consumer()
records = []
heart_rates = deque(maxlen=WINDOW_SIZE)
hr_list = []

if consumer is not None:
    while True:
        msg = consumer.poll(1.0)

        if msg is None:
            continue
        if msg.error():
            print(f"Consumer error: {msg.error()}")
            continue

        processed_message = process_message(msg)
        if processed_message is None:
            continue  

        date, time_received, heart_rate, activity_type, full_date = processed_message
        
        warning = 'No warning'
        if activity_type == 'resting':
            if heart_rate < 60:
                warning = 'Warning: Low heart rate'
            elif heart_rate > 90:
                warning = 'Warning: High heart rate'
        elif activity_type == 'walking':
            if heart_rate < 90:
                warning = 'Warning: Low heart rate'
            elif heart_rate > 130:
                warning = 'Warning: High heart rate'
        elif activity_type == 'running':
            if heart_rate > 145:
                warning = 'Warning: High heart rate'
        
        trend = ''
        if len(heart_rates) > 0:
            if heart_rate > heart_rates[-1]:
                trend = 'increasing'
            elif heart_rate < heart_rates[-1]:
                trend = 'decreasing'
            else:
                trend = 'steady'
        
        mean_hr = 0
        std_hr = 0
        z_score = 0
        hr_list.append(heart_rate)
        anomaly = "No"

        if len(hr_list) > 10:  
            mean_hr = np.mean(hr_list)
            std_hr = np.std(hr_list)
            z_score = (heart_rate - mean_hr) / std_hr
            if abs(z_score) > 2:
                anomaly = 'Yes'  
        else:
            z_score = None
            anomaly = None

        record = {
            'lp': len(records) + 1,
            'date': full_date,
            'time': time_received,
            'heart_rate': heart_rate,
            'activity_type': activity_type,
            'warning': warning,
            'trend': trend,
            'z-score': z_score,
            'anomaly': anomaly
        }
        records.append(record)
        heart_rates.append(heart_rate)

        
        if heart_rate > ALERT_THRESHOLDS.get(activity_type, 90):  
            print(f"ALERT: High heart rate detected! Date: {date}, Time: {time_received}, Activity: {activity_type}, Heart Rate: {heart_rate}")

        
        if len(heart_rates) == WINDOW_SIZE:
            avg_heart_rate = sum(heart_rates) / WINDOW_SIZE
            trend_over_window = "increasing" if heart_rates[-1] > heart_rates[0] else "decreasing"
            print(f"Date: {date}, Time: {time_received} - Average heart rate: {avg_heart_rate:.2f} - Trend: {trend_over_window}")

        if len(records) == 20:
            df = pd.DataFrame(records)
            print(df.to_string(index=False))

            filename = f"heartrate_data.xlsx"
            df.to_excel(filename, index=False)
            print(f"Table saved to file {filename}")
             
            add_line_chart_to_excel(filename)

            records = []

